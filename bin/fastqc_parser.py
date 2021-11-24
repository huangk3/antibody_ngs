#!/usr/bin/env python3

import glob 
import os, sys, argparse
from bs4 import BeautifulSoup as bs
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

parser = argparse.ArgumentParser(description='Optional app description')
parser.add_argument('-f', '--folder', required = True, type=str, help='A required integer positional argument')
parser.add_argument('-o', '--output_prefix', required = True, type=str, help='A required integer positional argument')

args = parser.parse_args()

html_folder=args.folder
htmls = glob.glob( html_folder + "/*.html")

#1-1-VH_combined_R1_fastqc.html

def extract_read_info (file_path):
  base_name=os.path.basename(file_path)
  read = os.path.splitext(base_name)[0].split('_')[3]
  return read

def extract_sample_id (file_path):
  base_name=os.path.basename(file_path)
  sample = os.path.splitext(base_name)[0].split('_')[0]    
  return sample

def per_base_summary (sample_list, df):
  fail_1 = 0
  fail_2 = 0
  other = 0
  for i in sample_list:
    n_failed = len(df[(df['Sample'] == i) & (df['PerBaseQual'] == 'FAIL')])
    if n_failed == 2 :
      fail_2 += 1 
    elif n_failed == 1 :
      fail_1 += 1
    elif n_failed == 0 :
      other += 1
    else:
      print("The number of failed reads is greater than 2!!")  
      sys.exit(1)
  df = pd.DataFrame([{'Category': 'Both Failed', 'Count': fail_2}, {'Category': 'One Failed', 'Count': fail_1}, \
          {'Category': 'Other', 'Count': other}, {'Category': 'Total', 'Count': fail_1 + fail_2 + other}]).sort_values(['Category'])
  return df
      
def write_excel(dataframe, sheet_name, wb):
  ws = wb.create_sheet(sheet_name)
  for r in dataframe_to_rows(dataframe, index=False, header=True):
    ws.append(r)  


info=[]

for html in htmls:
  f = open(html, encoding="utf8")
  soup = bs(f, features="lxml")
  sequence_count = soup.find("td", text="Total Sequences").find_next_sibling().text
  per_base_qual = soup.find("a", text="Per base sequence quality").find_previous().get("alt")[1:-1]
  per_tile_qual = soup.find("a", text="Per tile sequence quality").find_previous().get("alt")[1:-1]
  per_seq_qual = soup.find("a", text="Per sequence quality scores").find_previous().get("alt")[1:-1]

  ID = extract_sample_id(html)
  read = extract_read_info(html)
  info.append({'Sample': ID, 'Read': read, 'Count': int(sequence_count), "PerBaseQual": per_base_qual, "PerTileQual": per_tile_qual, "PerSeqQual": per_seq_qual})

df = pd.DataFrame(info).sort_values(['Sample', 'Read'])

samples = df['Sample'].drop_duplicates().to_list()


wb = Workbook()
ws = wb.active
ws.title = "fastqc_by_sample"
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)
ws['I2'] = "Warning"
ws['I3'] = "A warning will be issued if the lower quartile for any base is less than 10, or if the median for any base is less than 25."

ws['I5'] = "FAIL"
ws['I6'] = "This module will raise a failure if the lower quartile for any base is less than 5 or if the median for any base is less than 20."

per_base_summary_df = per_base_summary(samples, df)
write_excel(per_base_summary_df, "summary", wb)

wb.save(args.output_prefix+".xlsx")

